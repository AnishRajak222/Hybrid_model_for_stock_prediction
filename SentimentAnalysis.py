Excelfile=r"D:\Project\Observations\Observation.xlsx"
from openpyxl import Workbook, load_workbook
from LMSentiment import *
from datetime import datetime
import os
import json
os.environ['HF_HUB_DISABLE_SYMLINKS_WARNING'] = '1'
os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
from vaderSentiment.vaderSentiment import SentimentIntensityAnalyzer
from HeadlineScraper import *
import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification


DAILY_SENTIMENT_FILE = r"D:\Project\Observations\daily_sentiment.json"

# Load FinBERT model and tokenizer
def analyze_sentiment_FinBERT(headlines):
    sentiment_scores_sum = c = 0
    tokenizer = AutoTokenizer.from_pretrained("yiyanghkust/finbert-tone")
    model = AutoModelForSequenceClassification.from_pretrained("yiyanghkust/finbert-tone")
    for headline in headlines:
        # Tokenize and prepare input
        inputs = tokenizer(headline, return_tensors="pt", truncation=True, padding=True)
        # Run model and get logits
        with torch.no_grad():
            outputs = model(**inputs)
        # Apply softmax to get probabilities
        scores = torch.nn.functional.softmax(outputs.logits, dim=1)[0]
        # Extract probabilities for positive, neutral, and negative classes
        pos_score, neutral_score, neg_score = scores[0].item(), scores[1].item(), scores[2].item()
        # Calculate the normalized sentiment score in the range [-1, 1]
        sentiment_score = pos_score - neg_score
        sentiment_scores_sum += sentiment_score
        c += 1
    if c == 0:
        return 0
    return sentiment_scores_sum / c

def analyze_sentiment_Vader(headlines):
    analyzer = SentimentIntensityAnalyzer()
    scores = [analyzer.polarity_scores(headline)['compound'] for headline in headlines]
    average_score = sum(scores) / len(scores) if scores else 0
    return average_score

def load_daily_sentiment():
    """Load today's VADER and FinBERT scores if available."""
    if not os.path.exists(DAILY_SENTIMENT_FILE):
        return None
    today = datetime.now().strftime("%Y-%m-%d")
    with open(DAILY_SENTIMENT_FILE, "r") as file:
        data = json.load(file)
    return data.get(today)

def save_daily_sentiment(vader_score, finbert_score,lm_score):
    """Save today's VADER and FinBERT scores to a cache file."""
    today = datetime.now().strftime("%Y-%m-%d")
    if os.path.exists(DAILY_SENTIMENT_FILE):
        with open(DAILY_SENTIMENT_FILE, "r") as file:
            data = json.load(file)
    else:
        data = {}
    data[today] = {"vader": vader_score, "finbert": finbert_score,"lm": lm_score}
    with open(DAILY_SENTIMENT_FILE, "w") as file:
        json.dump(data, file, indent=4)
    save_to_Excel(vader_score, finbert_score,lm_score)#Delete this after obeservation is done

#Delete this after the observations are over
def save_to_Excel(vader_score, finbert_score,lm_score, file_path=Excelfile):
    """Save today's VADER and FinBERT scores to an Excel file."""
    today = datetime.now().strftime("%Y-%m-%d")
    sheet_name="Senti Scores"
    # Check if the file exists
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Date", "VADER Score", "FinBERT Score", "LM Score"])  # Add headers if new
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Date", "VADER Score", "FinBERT Score", "LM Score"])  # Add headers if new
    # Check if today's date already exists in the sheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == today:
            print(f"Entry for {today} already exists. Skipping.")
            wb.save(file_path)
            return
    # Append new data if today's entry doesn't exist
    ws.append([today, vader_score, finbert_score,lm_score])
    # Save workbook
    wb.save(file_path)
    print(f"Sentiment scores saved to '{sheet_name}' in {file_path}")
    # Append new data
    ws.append([today, vader_score, finbert_score,lm_score])
    # Save workbook
    wb.save(file_path)
    # print(f"Sentiment scores saved to '{sheet_name}' in {file_path}")


def CompanyName(symbol):
    inverse_stock_symbols = {
    "HCLTECH.NS": "HCL TECH",
    "INFY.NS": "INFOSYS",
    "TCS.NS": "TCS",
    "WIPRO.NS": "WIPRO",
    "TECHM.NS": "Tech Mahindra",
    "LTIM.NS": "LTIMindtree",
    "PERSISTENT.NS": "Persistent Systems",
    "OFSS.NS": "Oracle Financial Services",
    "POLICYBZR.NS": "Policy Bazar",
    "HEXT.NS": "Hexaware",
    "SASKEN.NS": "Sasken Tech",
    "QUICKHEAL.NS": "Quick Heal",
    "BSOFT.NS": "Birlasoft"}
    return inverse_stock_symbols[symbol]


def SentiWeight(symbol):
    file_path = r"C:\Users\Sahil\Desktop\Development\Machine Learning\Project\Observations\ga_weights.json"
    stock_name=CompanyName(symbol)
    try:
        with open(file_path, "r") as file:
            data = json.load(file)

            if stock_name in data:
                model = data[stock_name].get("Model Weights", {})
                sentiment = data[stock_name].get("Sentiment Weights", {})

                vader = sentiment.get("VADER Score", 0.0)
                finbert = sentiment.get("FinBERT Score", 0.0)
                lm = sentiment.get("LM Score", 0.0)

                return vader,finbert,lm
            else:
                raise ValueError(f"Stock '{stock_name}' not found.")
    
    except FileNotFoundError:
        raise FileNotFoundError("Weights file not found.")
    except json.JSONDecodeError:
        raise ValueError("Error decoding JSON file.")
    

import json
def SentimentScore(symbol):
    """Main function to calculate the sentiment score for a specific stock symbol."""
    # Load today's daily sentiment scores
    daily_sentiment = load_daily_sentiment()
    if not daily_sentiment:
        # Fetch headlines and calculate VADER and FinBERT scores if not cached
        print("Fetching headlines for the IT sector...")
        headlines = fetch_all_headlines()
        with open("HeadLines.json", "w") as file:
            json.dump(headlines, file)
        if not headlines:
            return 0
        vader_score = analyze_sentiment_Vader(headlines)
        finbert_score = analyze_sentiment_FinBERT(headlines)
        lm_score = LMSentimentScore(headlines)
        save_daily_sentiment(vader_score, finbert_score,lm_score)
        print ("vader_score", vader_score   )
        print ("FinBERT_score",finbert_score)
        print ("LMSentimentScore",lm_score  )
    else:
        vader_score = daily_sentiment["vader"]
        finbert_score = daily_sentiment["finbert"]
        lm_score =daily_sentiment["lm"]

    w1, w2 , w3 = SentiWeight(symbol)
    sentiment_score = (vader_score * w1 + finbert_score * w2 + lm_score*w3)
    return round(sentiment_score, 2)

if __name__ == "__main__":
    symbols = ["INFY.NS", "HCLTECH.NS", "TCS.NS", "WIPRO.NS"]
    # for symbol in symbols:
    #     score = SentimentScore(symbol)
    #     print(f"Final Sentiment Score for {symbol}: {score}")
    score = SentimentScore("INFY.NS")
